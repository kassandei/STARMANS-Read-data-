import serial
import time
from openpyxl import load_workbook
import os
from glob import glob

# Function to parse the data blocks
def parse_data_blocks(data_lines):
    blocks = []
    current_block = []
    for line in data_lines:
        if len(current_block) == 5:
            blocks.append(current_block)
            current_block = []
        current_block.append(line.decode().strip())
    if current_block:
        # If the last block has only 4 lines, duplicate the fourth line to make it 5
        while len(current_block) < 5:
            current_block.append(current_block[-1])
        blocks.append(current_block)
    return blocks

# Function to save data to the specified cells in the Excel file
def save_to_excel(blocks, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Define cell positions (adjust as per your Excel layout)
    cell_positions = [
        ('D10', 'E10'), ('K10', 'L10'),
        ('D22', 'E22'), ('K22', 'L22')
    ]
    
    for i, block in enumerate(blocks):
        if i >= len(cell_positions):
            break
        pos1, pos2 = cell_positions[i]
        start_row1 = int(pos1[1:])
        start_row2 = int(pos2[1:])
        col1 = pos1[0]
        col2 = pos2[0]
        for j, line in enumerate(block):
            if j >= 5:
                break
            data1, data2 = map(float, line.split(','))  # Convert to float (change to int if integers)
            ws[f'{col1}{start_row1 + j}'] = data1
            ws[f'{col2}{start_row2 + j}'] = data2
        
        # If the block has only 4 lines, copy the 4th line to the 5th line position
        if len(block) == 4:
            data1, data2 = map(float, block[3].split(','))
            ws[f'{col1}{start_row1 + 4}'] = data1
            ws[f'{col2}{start_row2 + 4}'] = data2
    
    wb.save(excel_path)
    print(f"Data saved to {excel_path}")

# Connect to the serial port (update port name for Windows)
ser = serial.Serial('/dev/tty.usbserial-1330', 9600, timeout=0)
data_lines = []

try:
    while True:
        try:
            data = ser.readline().strip()
            if data:
                print(data)  # Debugging: print received data
                data_lines.append(data)
                if len(data_lines) >= 20:  # Stop after reading 20 lines of data
                    break
            time.sleep(1)
        except serial.SerialTimeoutException:
            print('Data could not be read')  # Handle read timeout
except KeyboardInterrupt:
    pass

# Parse the data blocks
blocks = parse_data_blocks(data_lines)

# Get the current directory
current_directory = os.path.dirname(os.path.abspath(__file__))

# Find the first Excel file in the current directory
excel_files = glob(os.path.join(current_directory, '*.xlsx'))

# Check if any Excel file is found
if not excel_files:
    print("No Excel file found in the directory.")
else:
    # Use the first found Excel file
    excel_path = excel_files[0]
    print(f"Using Excel file: {excel_path}")

    # Save the parsed data to the Excel file
    save_to_excel(blocks, excel_path)
